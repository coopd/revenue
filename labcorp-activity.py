import numpy as np
import pandas as pd
import os
import glob
import re
import calendar
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

#path = "/Users/cdavis/Library/CloudStorage/Egnyte-indigobio/Shared/Production/Transaction-Reports/Monthlies"
path = "/Volumes/indigobio/Shared/Production/Transaction-Reports/Monthlies/"
csv_files = sorted(glob.glob(os.path.join(path, "*.csv")))


df = pd.DataFrame(columns=['site_name', 'processed_at', 'instrument_name', 'batch_name',
       'batch_id', 'assay_name', 'assay_id', 'sample_count',
       'determination_count', 'chromatogram_count', 'system'])

for file in csv_files:
    m = re.search(r'indigoascent4', file)
    if m:
        tf = pd.read_csv(file)
        tf['system'] = "S4"
    else:
        m = re.search(r'indigoarq1', file)
        if m:
            tf = pd.read_csv(file)
            tf['system'] = "ARQ"
        else:
            tf = pd.read_csv(file)
            tf['system'] = "S3"
            tf = tf.drop(columns=['total_db_determinations','total_db_size_in_mb'])
    df = pd.concat([df, tf], ignore_index=True)

df['date_time'] = pd.to_datetime(df['processed_at'], infer_datetime_format=True, utc=False)
df['year'] = [x.timetuple().tm_year for x in df['date_time']]
df['month'] = [x.timetuple().tm_mon for x in df['date_time']]
df['year_day'] = [x.timetuple().tm_yday for x in df['date_time']]

df = df.drop(columns=['determination_count'])

instrument_sn = pd.read_csv('labcorp_sn.csv')
license_loc = pd.read_csv("labcorp_license.csv")

# Make Excel Spreadsheet

wb = Workbook()

customer = 'LabCorp'

targets = {
    'Medtox': ['medtox-for', 'medtox-fortest', 'medtox-specialtytox', 'medtox-specialtytoxtest',
               'medtoxnew', 'medtoxtest'],
    'Burlington': ['labcorp', 'labcorptest', 'labcorptraining'],
    'RTP': ['labcorprtp', 'labcorprtptest', 'labcorprtp2', 'labcorprtp2test'],
    'Raritan': ["labcorpnewjerseyotstest"],
    'Covance_Indy': ['covanceindy', 'covanceindytest'],
    'Covance_Geneva': ['labcorpgvatest']
}

report_year = 2022
report_month = 12

for location in targets.keys():
    print(f'{location}: {targets[location]}')

    web_sites = targets[location]

    ws = wb.create_sheet(location)

    ws.append([f'{customer}: {location}, {calendar.month_name[report_month]}-{report_year}'])
    ws.append([' '])

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15

    loc_activity = (df.loc[df['site_name'].isin(web_sites) & df['year'].isin([report_year]) & df['month']
                    .isin([report_month])]
                    .groupby(['instrument_name', 'site_name', 'system'])[['chromatogram_count', 'sample_count']]
                    .agg(chromatogram_count=('chromatogram_count', np.sum),
                         sample_count=('sample_count', np.sum)))

    loc_activity = loc_activity.reset_index()
    #cols = ['instrument_name', 'site_name', 'system', 'chromatogram_count', 'sample_count']
    #loc_activity = loc_activity[cols]

    if len(loc_activity) != 0:
        loc_activity = pd.merge(loc_activity, instrument_sn, how='left', on='instrument_name')
        loc_activity['license_class'] = loc_activity['license_class'].fillna('prod')

        cols = ['instrument_name', 'serial_number', 'license_class', 'site_name', 'system', 'chromatogram_count', 'sample_count']
        loc_activity = loc_activity[cols]

        chrom_count = loc_activity['chromatogram_count'].sum()
        sample_count = loc_activity['sample_count'].sum()
        active_prod = len(loc_activity[loc_activity['license_class'] == 'prod']['instrument_name'].unique())
        active_dev = len(loc_activity[loc_activity['license_class'] == 'dev']['instrument_name'].unique())
    else:
        loc_activity = pd.DataFrame(columns=['instrument_name', 'serial_number', 'license_class',
                                             'site_name', 'system', 'chromatogram_count', 'sample_count'])
        chrom_count = 0
        sample_count = 0
        active_prod = 0
        active_dev = 0

    owned_prod = license_loc[license_loc['location'] == location]['prod'].values[0]
    owned_dev = license_loc[license_loc['location'] == location]['dev'].values[0]

    for r in dataframe_to_rows(loc_activity, index=False, header=True):
        ws.append(r)

    ws.append({'F': chrom_count, 'G': sample_count})
    ws.append([' '])
    ws.append({'A': 'Prod Licenses', 'B': owned_prod})
    ws.append({'A': 'Prod Active', 'B': active_prod})
    ws.append([' '])
    ws.append({'A': 'Dev Licenses', 'B': owned_dev})
    ws.append({'A': 'Dev Active', 'B': active_dev})

# Get rid of the default empty sheet when workbook is created

if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Save Excel spreadsheet

report_file = f'{report_year}_{str(report_month).zfill(2)}_activity.xlsx'
wb.save(report_file)